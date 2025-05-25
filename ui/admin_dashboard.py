import sys
import os
from pathlib import Path
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                           QPushButton, QTableWidget, QTableWidgetItem, 
                           QTabWidget, QMessageBox, QHeaderView, QFrame,
                           QSplitter, QComboBox, QLineEdit, QFormLayout,
                           QScrollArea, QGroupBox, QTextEdit, QDialog,
                           QFileDialog, QGridLayout, QListWidget, QListWidgetItem,
                           QStackedWidget, QButtonGroup, QRadioButton, QProgressBar,
                           QCheckBox)
from PyQt5.QtCore import Qt, QSize, QUrl
from PyQt5.QtGui import QIcon, QColor, QPixmap
from PyQt5.QtGui import QDesktopServices
import openpyxl
import logging
from datetime import datetime
import subprocess

# Add the parent directory to the path to allow imports
parent_dir = str(Path(__file__).resolve().parent.parent)
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

from db.database import (
    get_user_name, get_reports_by_status, get_all_reports, 
    update_report_status, get_report, add_approval_log, get_user_by_id, add_template, 
    get_active_template, get_all_users, delete_user, update_user,
    get_all_templates, delete_template, is_template_active,
    get_leader_approval
)
from utils.excel_handler import ExcelHandler
from utils.auth import Auth
from utils.notifications import send_report_notification
from email_sender import EmailSender
from pdf.pdf_generator import PDFGenerator

class UserManagementDialog(QDialog):
    """Dialog for adding/editing users"""
    def __init__(self, parent=None, user=None):
        super().__init__(parent)
        self.user = user
        
        if user:
            self.setWindowTitle("Edit User")
        else:
            self.setWindowTitle("Add New User")
        
        self.resize(400, 400)  # Increased height for new fields
        
        layout = QVBoxLayout(self)
        
        form_layout = QFormLayout()
        
        # Username field
        self.username_input = QLineEdit()
        if user:
            self.username_input.setText(user.get('username', ''))
        form_layout.addRow("Username:", self.username_input)
        
        # Password field
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        form_layout.addRow("Password:", self.password_input)
        
        # Email field
        self.email_input = QLineEdit()
        if user:
            self.email_input.setText(user.get('email', ''))
        form_layout.addRow("Email:", self.email_input)
        
        # Employee Code field
        self.emp_code_input = QLineEdit()
        if user:
            self.emp_code_input.setText(user.get('emp_code', ''))
        form_layout.addRow("Employee Code:", self.emp_code_input)
        
        # Designation field
        self.designation_input = QLineEdit()
        if user:
            self.designation_input.setText(user.get('designation', ''))
        form_layout.addRow("Designation:", self.designation_input)
        
        # Digital signature info
        signature_info = QLabel("Employee code and designation will be used for digital signature")
        signature_info.setStyleSheet("font-style: italic; color: #3498db;")
        form_layout.addRow("", signature_info)
        
        # Role dropdown
        self.role_dropdown = QComboBox()
        self.role_dropdown.addItem("User", "user")
        self.role_dropdown.addItem("Unit Leader", "unit_leader")
        self.role_dropdown.addItem("Admin", "admin")
        
        if user:
            # Set the current role
            index = self.role_dropdown.findData(user.get('role', 'user'))
            if index >= 0:
                self.role_dropdown.setCurrentIndex(index)
        
        form_layout.addRow("Role:", self.role_dropdown)
        
        layout.addLayout(form_layout)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_button)
        
        save_button = QPushButton("Save")
        save_button.clicked.connect(self.accept)
        buttons_layout.addWidget(save_button)
        
        layout.addLayout(buttons_layout)
    
    def get_user_data(self):
        """Get the user data from the form"""
        return {
            'username': self.username_input.text(),
            'password': self.password_input.text(),
            'email': self.email_input.text(),
            'emp_code': self.emp_code_input.text(),
            'designation': self.designation_input.text(),
            'role': self.role_dropdown.currentData()
        }

class TemplatePreviewDialog(QDialog):
    """Dialog to preview Excel template structure"""
    def __init__(self, template_path, parent=None):
        super().__init__(parent)
        self.template_path = template_path
        self.setWindowTitle("Template Preview")
        self.resize(800, 600)
        
        layout = QVBoxLayout(self)
        
        # Add template info
        info_label = QLabel(f"Template: {os.path.basename(template_path)}")
        layout.addWidget(info_label)
        
        # Create tabs for sheet preview
        tabs = QTabWidget()
        layout.addWidget(tabs)
        
        try:
            # Load the Excel file
            workbook = openpyxl.load_workbook(template_path, data_only=True)
            
            # Add each sheet as a tab
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Create a table for this sheet
                sheet_table = QTableWidget()
                tabs.addTab(sheet_table, sheet_name)
                
                # Configure the table
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Limit preview to first 100 rows for performance
                preview_rows = min(max_row, 100)
                
                sheet_table.setRowCount(preview_rows)
                sheet_table.setColumnCount(max_col)
                
                # Set headers from first row
                for col in range(1, max_col + 1):
                    header_cell = sheet.cell(row=1, column=col)
                    header_value = str(header_cell.value or "")
                    sheet_table.setHorizontalHeaderItem(col-1, QTableWidgetItem(header_value))
                
                # Load data into table
                for row in range(2, preview_rows + 1):  # Skip header row
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        value = str(cell.value or "")
                        item = QTableWidgetItem(value)
                        sheet_table.setItem(row-2, col-1, item)  # Adjust for 0-based indexing
                
                # Auto-resize columns for better view
                sheet_table.resizeColumnsToContents()
            
        except Exception as e:
            error_label = QLabel(f"Error loading template: {str(e)}")
            layout.addWidget(error_label)
        
        # Add close button
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button)

class ReportDetailDialog(QDialog):
    """Dialog to display detailed report information and approval history"""
    
    def __init__(self, report, parent=None):
        super().__init__(parent)
        self.report = report
        self.user = parent.user if parent else None
        self.setup_ui()
    
    def setup_ui(self):
        """Set up the UI elements for the dialog"""
        self.setWindowTitle(f"Report Details: {self.report['title']}")
        self.setMinimumSize(800, 600)
        
        layout = QVBoxLayout(self)
        
        # Header
        header = QLabel(f"<h2>{self.report['title']}</h2>")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        # Create tab widget for different sections
        tab_widget = QTabWidget()
        layout.addWidget(tab_widget)
        
        # Basic Info Tab
        info_tab = QWidget()
        info_layout = QVBoxLayout(info_tab)
        self.setup_info_tab(info_layout)
        tab_widget.addTab(info_tab, "Report Information")
        
        # Content Tab
        content_tab = QWidget()
        content_layout = QVBoxLayout(content_tab)
        self.setup_content_tab(content_layout)
        tab_widget.addTab(content_tab, "Content")
        
        # Approval History Tab
        history_tab = QWidget()
        history_layout = QVBoxLayout(history_tab)
        self.setup_history_tab(history_layout)
        tab_widget.addTab(history_tab, "Approval History")
        
        # Bottom buttons
        button_layout = QHBoxLayout()
        
        # Only show action buttons if the user is an admin
        if self.user and self.user.get('role') == 'admin':
            if self.report['status'] == 'pending_approval' or self.report['status'] == 'leader_approved':
                approve_btn = QPushButton("Approve Report")
                approve_btn.setStyleSheet("background-color: #27ae60; color: white;")
                approve_btn.clicked.connect(self.approve_report)
                button_layout.addWidget(approve_btn)
                
                sendback_btn = QPushButton("Send Back for Revision")
                sendback_btn.setStyleSheet("background-color: #e67e22; color: white;")
                sendback_btn.clicked.connect(self.send_back_report)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
    
    def setup_info_tab(self, layout):
        """Set up the info tab with the basic report information"""
        # Report info in a form layout
        form_group = QGroupBox("Basic Information")
        form_layout = QFormLayout(form_group)
        
        # Add report info fields
        report_id = self.report.get('report_id', 'N/A')
        form_layout.addRow("Report ID:", QLabel(str(report_id)))
        
        user_id = self.report.get('user_id')
        author_name = get_user_name(user_id) if user_id else "Unknown"
        form_layout.addRow("Author:", QLabel(author_name))
        
        if 'created_at' in self.report and self.report['created_at']:
            try:
                created_date = self.report['created_at'].strftime("%Y-%m-%d %H:%M")
            except (AttributeError, TypeError):
                created_date = str(self.report['created_at'])
            form_layout.addRow("Created:", QLabel(created_date))
        
        if 'last_modified_at' in self.report and self.report['last_modified_at']:
            try:
                updated_date = self.report['last_modified_at'].strftime("%Y-%m-%d %H:%M")
            except (AttributeError, TypeError):
                updated_date = str(self.report['last_modified_at'])
            form_layout.addRow("Last Updated:", QLabel(updated_date))
        
        # Add user's digital signature if available
        if 'fields' in self.report and 'user_signature' in self.report['fields']:
            signature = self.report['fields']['user_signature']
            signature_label = QLabel(signature)
            signature_label.setStyleSheet("font-weight: bold; color: #16a085;")
            form_layout.addRow("User Signature:", signature_label)
        
        # Status with colored indicator
        status_layout = QHBoxLayout()
        status_text = self.report.get('status', 'Unknown').replace('_', ' ').title()
        status_label = QLabel(status_text)
        
        # Add color indicator based on status
        if self.report.get('status') == 'approved':
            status_label.setStyleSheet("color: green; font-weight: bold;")
        elif self.report.get('status') == 'needs_revision':
            status_label.setStyleSheet("color: orange; font-weight: bold;")
        elif self.report.get('status') == 'leader_approved':
            status_label.setStyleSheet("color: blue; font-weight: bold;")
        else:
            status_label.setStyleSheet("color: gray; font-weight: bold;")
        
        status_layout.addWidget(status_label)
        form_layout.addRow("Status:", status_layout)
        
        layout.addWidget(form_group)
        
        # Add Excel file section if available
        if 'fields' in self.report and 'excel_file_path' in self.report['fields']:
            excel_path = self.report['fields']['excel_file_path']
            if excel_path and os.path.exists(excel_path):
                excel_group = QGroupBox("Excel Report File")
                excel_layout = QVBoxLayout(excel_group)
                
                file_info = QLabel(f"File: {os.path.basename(excel_path)}")
                excel_layout.addWidget(file_info)
                
                open_btn = QPushButton("Open Excel File")
                open_btn.setStyleSheet("background-color: #3498db; color: white;")
                open_btn.clicked.connect(lambda: self.open_excel_file(excel_path))
                excel_layout.addWidget(open_btn)
                
                layout.addWidget(excel_group)
        
        # Add metadata if any
        if 'metadata' in self.report and self.report['metadata']:
            metadata_group = QGroupBox("Additional Metadata")
            metadata_layout = QFormLayout(metadata_group)
            
            for key, value in self.report['metadata'].items():
                if value:  # Only display non-empty values
                    metadata_layout.addRow(f"{key.replace('_', ' ').title()}:", QLabel(str(value)))
            
            layout.addWidget(metadata_group)
    
    def open_excel_file(self, file_path):
        """Open an Excel file in the default application"""
        try:
            if sys.platform == 'win32':
                os.startfile(file_path)
            elif sys.platform == 'darwin':  # macOS
                subprocess.run(['open', file_path])
            else:  # Linux
                subprocess.run(['xdg-open', file_path])
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to open Excel file: {str(e)}")
    
    def setup_content_tab(self, layout):
        """Set up the tab showing the report content"""
        content_display = QTextEdit()
        content_display.setReadOnly(True)
        
        # Set the content, handling cases where it might be missing
        if 'content' in self.report and self.report['content']:
            content_display.setHtml(self.report['content'])
        else:
            content_display.setPlainText("No content available for this report.")
        
        layout.addWidget(QLabel("<h3>Report Content</h3>"))
        layout.addWidget(content_display)
        
        # Add attachments section if there are any
        if 'attachments' in self.report and self.report['attachments']:
            attachments_group = QGroupBox("Attachments")
            attachments_layout = QVBoxLayout(attachments_group)
            
            for attachment in self.report['attachments']:
                attachment_layout = QHBoxLayout()
                
                filename = attachment.get('filename', 'Unnamed attachment')
                attachment_label = QLabel(filename)
                attachment_layout.addWidget(attachment_label)
                
                view_btn = QPushButton("View")
                view_btn.clicked.connect(lambda _, a=attachment: self.view_attachment(a))
                attachment_layout.addWidget(view_btn)
                
                attachments_layout.addLayout(attachment_layout)
            
            layout.addWidget(attachments_group)
    
    def setup_history_tab(self, layout):
        """Set up the tab showing the approval history"""
        layout.addWidget(QLabel("<h3>Approval Timeline</h3>"))
        
        # Create a timeline widget
        timeline_widget = QWidget()
        timeline_layout = QVBoxLayout(timeline_widget)
        timeline_layout.setSpacing(15)
        
        # Get approval logs or use empty list if not available
        approval_logs = []
        if 'approval_logs' in self.report and self.report['approval_logs']:
            try:
                # Sort logs by timestamp
                approval_logs = sorted(
                    self.report['approval_logs'], 
                    key=lambda x: x['timestamp'] if 'timestamp' in x else datetime.now()
                )
            except Exception as e:
                # If sorting fails, still try to display them unsorted
                approval_logs = self.report['approval_logs']
                print(f"Error sorting approval logs: {e}")
        
        if not approval_logs:
            # Show placeholder if no logs
            no_logs_label = QLabel("No approval history available for this report.")
            no_logs_label.setAlignment(Qt.AlignCenter)
            no_logs_label.setStyleSheet("color: #7f8c8d; font-style: italic;")
            timeline_layout.addWidget(no_logs_label)
        
        for i, log in enumerate(approval_logs):
            log_widget = QGroupBox()
            log_layout = QVBoxLayout(log_widget)
            
            # Header with timestamp and action
            action_text = log.get('action', 'unknown').replace('_', ' ').title()
            header_layout = QHBoxLayout()
            
            # Add appropriate icon and color based on action
            if "approve" in log.get('action', ''):
                status_color = "#27ae60"  # Green for approvals
                action_prefix = "Approved"
            elif "send_back" in log.get('action', ''):
                status_color = "#e67e22"  # Orange for send backs
                action_prefix = "Sent Back"
            elif "submit" in log.get('action', ''):
                status_color = "#3498db"  # Blue for submissions
                action_prefix = "Submitted"
            else:
                status_color = "#7f8c8d"  # Grey for other actions
                action_prefix = action_text
            
            # Display timestamp 
            timestamp_text = "Unknown date"
            if 'timestamp' in log and log['timestamp']:
                try:
                    timestamp_text = log['timestamp'].strftime('%Y-%m-%d %H:%M')
                except (AttributeError, TypeError):
                    timestamp_text = str(log['timestamp'])
            
            timestamp = QLabel(f"<i>{timestamp_text}</i>")
            timestamp.setStyleSheet(f"color: {status_color};")
            header_layout.addWidget(timestamp)
            
            # Action label with user who performed it
            action_by_name = "Unknown" 
            if 'action_by' in log:
                action_by_name = get_user_name(log['action_by'])
            elif 'actor_name' in log:
                action_by_name = log['actor_name']
                
            action_label = QLabel(f"<b>{action_prefix} by {action_by_name}</b>")
            action_label.setStyleSheet(f"color: {status_color};")
            header_layout.addWidget(action_label)
            
            header_layout.addStretch()
            log_layout.addLayout(header_layout)
            
            # Comments
            if 'comments' in log and log['comments']:
                comments = QLabel(log['comments'])
                comments.setWordWrap(True)
                comments.setStyleSheet("margin-left: 20px;")
                log_layout.addWidget(comments)
            
            timeline_layout.addWidget(log_widget)
            
            # Add connector line between items (except for the last one)
            if i < len(approval_logs) - 1:
                line = QFrame()
                line.setFrameShape(QFrame.VLine)
                line.setFrameShadow(QFrame.Sunken)
                line.setFixedHeight(20)
                line_layout = QHBoxLayout()
                line_layout.addStretch()
                line_layout.addWidget(line)
                line_layout.addStretch()
                timeline_layout.addLayout(line_layout)
        
        # Put the timeline in a scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(timeline_widget)
        layout.addWidget(scroll_area)
    
    def view_attachment(self, attachment):
        """Open an attachment from the report"""
        try:
            if not attachment or 'file_path' not in attachment:
                QMessageBox.warning(self, "Error", "Invalid attachment data")
                return
                
            file_path = attachment['file_path']
            if not file_path or not os.path.exists(file_path):
                QMessageBox.warning(self, "Error", f"Attachment file not found at {file_path}")
                return
                
            url = QUrl.fromLocalFile(file_path)
            QDesktopServices.openUrl(url)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not open attachment: {str(e)}")
    
    def approve_report(self):
        """Call the parent's approve report method"""
        self.accept()
        # Check if parent has approve_report method
        if hasattr(self.parent(), 'approve_report'):
            self.parent().approve_report(self.report['id'])
    
    def send_back_report(self):
        """Call the parent's send back report method"""
        self.accept()
        # Check if parent has send_back_report method 
        if hasattr(self.parent(), 'send_back_report'):
            self.parent().send_back_report(self.report['id'])

class AdminDashboard(QWidget):
    def __init__(self, user):
        super().__init__()
        self.user = user
        print(f"DEBUG: AdminDashboard initialized with user: {self.user}")
        self.email_sender = EmailSender()
        self.active_template = None
        self.init_ui()
        self.refresh_data()
        
    def init_ui(self):
        """Initialize the UI"""
        # Create main layout
        main_layout = QVBoxLayout(self)

        # Create tab widget for different sections
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)

        # Dashboard tab
        self.dashboard_tab = QWidget()
        self.tab_widget.addTab(self.dashboard_tab, "Dashboard")
        self.setup_dashboard_tab()

        # Users tab
        self.users_tab = QWidget()
        self.tab_widget.addTab(self.users_tab, "Users")
        self.setup_users_tab()

        # Templates tab
        self.templates_tab = QWidget()
        self.tab_widget.addTab(self.templates_tab, "Templates")
        self.setup_templates_tab()

        # Approvals tab
        self.approvals_tab = QWidget()
        self.tab_widget.addTab(self.approvals_tab, "Approvals")
        self.setup_approvals_tab()

        # Reports tab
        self.reports_tab = QWidget()
        self.tab_widget.addTab(self.reports_tab, "All Reports")
        self.setup_reports_tab()

        # Settings tab
        self.settings_tab = QWidget()
        self.tab_widget.addTab(self.settings_tab, "Settings")
        self.setup_settings_tab()

    def setup_settings_tab(self):
        """Set up the settings tab for admin preferences"""
        from ui.settings_dialog import SettingsDialog
        layout = QVBoxLayout(self.settings_tab)
        info_label = QLabel("Configure email notification and workflow preferences below.")
        info_label.setStyleSheet("font-size: 14px; color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(info_label)
        open_settings_btn = QPushButton("Open Settings")
        open_settings_btn.setStyleSheet("background-color: #2980b9; color: white; padding: 10px; border-radius: 4px;")
        open_settings_btn.clicked.connect(self.open_settings_dialog)
        layout.addWidget(open_settings_btn)
        layout.addStretch()

    def open_settings_dialog(self):
        from ui.settings_dialog import SettingsDialog
        dlg = SettingsDialog(self)
        dlg.exec_()

    def setup_dashboard_tab(self):
        """Set up the dashboard tab with summary info"""
        layout = QVBoxLayout(self.dashboard_tab)
        
        # Add welcome section
        welcome_frame = QFrame()
        welcome_frame.setStyleSheet("""
            QFrame {
                background-color: #ecf0f1;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        welcome_layout = QVBoxLayout(welcome_frame)
        
        welcome_label = QLabel(f"Welcome, {self.user['username']}!")
        welcome_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        welcome_layout.addWidget(welcome_label)
        
        instructions = QLabel("As an administrator, you can manage users, templates, and provide final approval for reports.")
        instructions.setWordWrap(True)
        welcome_layout.addWidget(instructions)
        
        layout.addWidget(welcome_frame)
        
        # Add stats section
        stats_frame = QFrame()
        stats_frame.setStyleSheet("""
            QFrame {
                background-color: #f9f9f9;
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        stats_layout = QHBoxLayout(stats_frame)
        
        # Stats will be populated in refresh_data
        self.pending_count_label = QLabel("Pending Approvals: 0")
        self.pending_count_label.setStyleSheet("font-size: 16px; color: #e74c3c;")
        stats_layout.addWidget(self.pending_count_label)
        
        self.users_count_label = QLabel("Users: 0")
        self.users_count_label.setStyleSheet("font-size: 16px; color: #3498db;")
        stats_layout.addWidget(self.users_count_label)
        
        self.approved_count_label = QLabel("Approved Reports: 0")
        self.approved_count_label.setStyleSheet("font-size: 16px; color: #2ecc71;")
        stats_layout.addWidget(self.approved_count_label)
        
        layout.addWidget(stats_frame)
        
        # Quick actions
        actions_label = QLabel("Quick Actions")
        actions_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-top: 15px;")
        layout.addWidget(actions_label)
        
        actions_frame = QFrame()
        actions_layout = QHBoxLayout(actions_frame)
        
        manage_users_btn = QPushButton("Manage Users")
        manage_users_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        manage_users_btn.clicked.connect(lambda: self.tab_widget.setCurrentIndex(1))
        actions_layout.addWidget(manage_users_btn)
        
        template_btn = QPushButton("Manage Templates")
        template_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        template_btn.clicked.connect(lambda: self.tab_widget.setCurrentIndex(2))
        actions_layout.addWidget(template_btn)
        
        approve_reports_btn = QPushButton("Review Pending Approvals")
        approve_reports_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 4px;
                font-weight: bold;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        approve_reports_btn.clicked.connect(lambda: self.tab_widget.setCurrentIndex(3))
        actions_layout.addWidget(approve_reports_btn)
        
        settings_btn = QPushButton("Settings")
        settings_btn.setStyleSheet("background-color: #2980b9; color: white; padding: 8px 16px; border-radius: 4px;")
        settings_btn.setMaximumWidth(120)
        settings_btn.clicked.connect(self.open_settings_dialog)
        actions_layout.addWidget(settings_btn)
        
        layout.addWidget(actions_frame)
        
        # Recent pending approvals section
        pending_label = QLabel("Pending Approvals")
        pending_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-top: 15px;")
        layout.addWidget(pending_label)
        
        self.pending_reports_table = QTableWidget()
        self.pending_reports_table.setColumnCount(5)
        self.pending_reports_table.setHorizontalHeaderLabels(
            ["ID", "Title", "Created by", "Unit Leader", "Actions"]
        )
        self.pending_reports_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.pending_reports_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.pending_reports_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.pending_reports_table)
        
        # Add spacer at the bottom
        layout.addStretch()
    
    def setup_users_tab(self):
        """Set up the user management tab"""
        layout = QVBoxLayout(self.users_tab)
        
        # Add instructions
        instructions = QLabel("Manage user accounts. Add, edit, or remove users.")
        instructions.setWordWrap(True)
        layout.addWidget(instructions)
        
        # Add actions
        actions_layout = QHBoxLayout()
        
        add_user_btn = QPushButton("Add New User")
        add_user_btn.setStyleSheet("background-color: #2ecc71; color: white;")
        add_user_btn.clicked.connect(self.add_user)
        actions_layout.addWidget(add_user_btn)
        
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh_users_table)
        actions_layout.addWidget(refresh_btn)
        
        actions_layout.addStretch()
        
        layout.addLayout(actions_layout)
        
        # Create users table
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(7)
        self.users_table.setHorizontalHeaderLabels(
            ["ID", "Username", "Role", "Email", "Emp Code", "Designation", "Actions"]
        )
        self.users_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.users_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.users_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.users_table)
    
    def setup_templates_tab(self):
        """Set up the templates management tab"""
        templates_layout = QVBoxLayout()
        
        # Header and buttons area
        header_layout = QHBoxLayout()
        
        # Title
        title = QLabel("<h2>Template Management</h2>")
        title.setStyleSheet("color: #2c3e50;")
        header_layout.addWidget(title)
        
        # Spacer
        header_layout.addStretch()
        
        # Upload button
        upload_btn = QPushButton("Upload New Template")
        upload_btn.setStyleSheet("""
            background-color: #2980b9;
            color: white;
            padding: 8px 15px;
            font-weight: bold;
        """)
        upload_btn.setIcon(QIcon(os.path.join('icons', 'upload.png')))
        upload_btn.clicked.connect(self.upload_template)
        header_layout.addWidget(upload_btn)
        
        templates_layout.addLayout(header_layout)
        
        # Templates table
        self.templates_table = QTableWidget()
        self.templates_table.setAlternatingRowColors(True)
        self.templates_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.templates_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.templates_table.setSelectionMode(QTableWidget.SingleSelection)
        self.templates_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #ecf0f1;
                gridline-color: #bdc3c7;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 5px;
                font-weight: bold;
            }
        """)
        
        # Set up columns
        self.templates_table.setColumnCount(5)
        self.templates_table.setHorizontalHeaderLabels(
            ["ID", "Name", "Uploaded By", "Date", "Actions"]
        )
        self.templates_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        
        templates_layout.addWidget(self.templates_table)
        
        # Set tab layout
        self.templates_tab.setLayout(templates_layout)
    
    def setup_approvals_tab(self):
        """Set up the pending approvals tab"""
        layout = QVBoxLayout(self.approvals_tab)
        
        # Add instructions
        instructions = QLabel(
            "Review and approve reports that have been approved by Unit Leaders."
        )
        instructions.setWordWrap(True)
        layout.addWidget(instructions)
        
        # Create pending approvals table
        self.approvals_table = QTableWidget()
        self.approvals_table.setColumnCount(6)
        self.approvals_table.setHorizontalHeaderLabels(
            ["ID", "Title", "Created by", "Submitted on", "Approved by Unit Leader", "Actions"]
        )
        self.approvals_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.approvals_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.approvals_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.approvals_table)
        
        # Add refresh button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh_data)
        layout.addWidget(refresh_btn, alignment=Qt.AlignRight)
    
    def setup_reports_tab(self):
        """Set up the all reports tab"""
        layout = QVBoxLayout(self.reports_tab)
        
        # Create filter controls
        filter_frame = QFrame()
        filter_layout = QHBoxLayout(filter_frame)
        
        filter_label = QLabel("Filter:")
        filter_layout.addWidget(filter_label)
        
        self.status_filter = QComboBox()
        self.status_filter.addItem("All Reports", "all")
        self.status_filter.addItem("Submitted", "submitted")
        self.status_filter.addItem("Approved by Unit Leader", "approved_leader")
        self.status_filter.addItem("Approved Final", "approved_admin")
        self.status_filter.addItem("Needs Revision", "needs_revision")
        self.status_filter.currentIndexChanged.connect(self.refresh_all_reports_table)
        filter_layout.addWidget(self.status_filter)
        
        search_label = QLabel("Search:")
        filter_layout.addWidget(search_label)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search in report titles...")
        self.search_input.textChanged.connect(self.refresh_all_reports_table)
        filter_layout.addWidget(self.search_input)
        
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh_data)
        filter_layout.addWidget(refresh_btn)
        
        layout.addWidget(filter_frame)
        
        # Create reports table
        self.all_reports_table = QTableWidget()
        self.all_reports_table.setColumnCount(6)
        self.all_reports_table.setHorizontalHeaderLabels(
            ["ID", "Title", "Created by", "Submitted on", "Status", "Actions"]
        )
        self.all_reports_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.all_reports_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.all_reports_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.all_reports_table) 

    def refresh_data(self):
        """Refresh all data on the dashboard"""
        # Refresh tables
        self.refresh_pending_tables()
        self.refresh_users_table()
        self.refresh_all_reports_table()
        self.refresh_dashboard_stats()
        self.refresh_templates_table()
    
    def refresh_pending_tables(self):
        """Refresh the pending approvals tables"""
        # Get pending approval reports (approved by unit leader)
        pending_reports = get_reports_by_status('approved_leader')
        
        # Update dashboard pending table
        self.pending_reports_table.setRowCount(0)
        
        # Update detailed pending table
        self.approvals_table.setRowCount(0)
        
        for row, report in enumerate(pending_reports):
            # Add to both tables
            self.pending_reports_table.insertRow(row)
            self.approvals_table.insertRow(row)
            
            # ID
            id_item = QTableWidgetItem(str(report['report_id']))
            self.pending_reports_table.setItem(row, 0, id_item)
            self.approvals_table.setItem(row, 0, QTableWidgetItem(str(report['report_id'])))
            
            # Title
            title_item = QTableWidgetItem(report['title'])
            self.pending_reports_table.setItem(row, 1, title_item)
            self.approvals_table.setItem(row, 1, QTableWidgetItem(report['title']))
            
            # Created by
            creator_item = QTableWidgetItem(report['creator_name'])
            self.pending_reports_table.setItem(row, 2, creator_item)
            self.approvals_table.setItem(row, 2, QTableWidgetItem(report['creator_name']))
            
            # Get approval info for the unit leader
            unit_leader_name = "Unknown"
            submitted_date = report['created_at']
            
            # Get the full report to see approval logs
            full_report = get_report(report['report_id'])
            if full_report and 'approval_logs' in full_report:
                for log in full_report['approval_logs']:
                    if log['action'] == 'approve_leader':
                        unit_leader_name = log['actor_name']
                        submitted_date = log['timestamp']
                        break
            
            # Unit Leader
            self.pending_reports_table.setItem(row, 3, QTableWidgetItem(unit_leader_name))
            
            # Submitted on (for approvals tab)
            self.approvals_table.setItem(row, 3, QTableWidgetItem(str(report['created_at'])))
            
            # Approved by Unit Leader date (for approvals tab)
            self.approvals_table.setItem(row, 4, QTableWidgetItem(str(submitted_date)))
            
            # Actions
            for table_idx, table in enumerate([self.pending_reports_table, self.approvals_table]):
                actions_cell = QWidget()
                actions_layout = QHBoxLayout(actions_cell)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                
                # View button
                view_btn = QPushButton("View")
                view_btn.setStyleSheet("background-color: #3498db; color: white;")
                view_btn.clicked.connect(lambda checked, r=report['report_id']: self.view_report(r))
                actions_layout.addWidget(view_btn)
                
                # Approve button
                approve_btn = QPushButton("Approve")
                approve_btn.setStyleSheet("background-color: #2ecc71; color: white;")
                approve_btn.clicked.connect(lambda checked, r=report['report_id']: self.approve_report(r))
                actions_layout.addWidget(approve_btn)
                
                # Send Back button
                reject_btn = QPushButton("Send Back")
                reject_btn.setStyleSheet("background-color: #e74c3c; color: white;")
                reject_btn.clicked.connect(lambda checked, r=report['report_id']: self.send_back_report(r))
                actions_layout.addWidget(reject_btn)
                
                if table_idx == 0:  # Dashboard table
                    table.setCellWidget(row, 4, actions_cell)
                else:  # Detailed table
                    table.setCellWidget(row, 5, actions_cell)
    
    def refresh_users_table(self):
        """Refresh the users table"""
        users = get_all_users()
        
        # Clear the table
        self.users_table.setRowCount(0)
        
        # Update columns to include emp code and designation
        self.users_table.setColumnCount(7)
        self.users_table.setHorizontalHeaderLabels(
            ["ID", "Username", "Role", "Email", "Emp Code", "Designation", "Actions"]
        )
        
        for row, user in enumerate(users):
            try:
                self.users_table.insertRow(row)
                
                # ID
                id_item = QTableWidgetItem(str(user.get('id', '')))
                self.users_table.setItem(row, 0, id_item)
                
                # Username
                username_item = QTableWidgetItem(user.get('username', ''))
                self.users_table.setItem(row, 1, username_item)
                
                # Role
                role_item = QTableWidgetItem(user.get('role', ''))
                self.users_table.setItem(row, 2, role_item)
                
                # Email
                email_item = QTableWidgetItem(user.get('email', ''))
                self.users_table.setItem(row, 3, email_item)
                
                # Employee Code
                emp_code_item = QTableWidgetItem(user.get('emp_code', '') or '')
                self.users_table.setItem(row, 4, emp_code_item)
                
                # Designation
                designation_item = QTableWidgetItem(user.get('designation', '') or '')
                self.users_table.setItem(row, 5, designation_item)
                
                # Actions
                actions_cell = QWidget()
                actions_layout = QHBoxLayout(actions_cell)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                
                # Edit button
                edit_btn = QPushButton("Edit")
                edit_btn.setStyleSheet("background-color: #f39c12; color: white;")
                edit_btn.clicked.connect(lambda checked, u=user: self.edit_user(u))
                actions_layout.addWidget(edit_btn)
                
                # Don't allow deleting the current user or the default admin
                if user.get('id') != self.user.get('id') and user.get('username') != 'admin':
                    delete_btn = QPushButton("Delete")
                    delete_btn.setStyleSheet("background-color: #e74c3c; color: white;")
                    delete_btn.clicked.connect(lambda checked, u=user: self.delete_user(u))
                    actions_layout.addWidget(delete_btn)
                
                self.users_table.setCellWidget(row, 6, actions_cell)
            except Exception as e:
                import logging
                logging.error(f"Error processing user row: {e}")
                continue
        
        # Resize columns for better readability
        self.users_table.resizeColumnsToContents()
    
    def refresh_all_reports_table(self):
        """Refresh the all reports table"""
        # Get filter values
        status_filter = self.status_filter.currentData()
        search_text = self.search_input.text().lower()
        
        # Get reports based on filter
        if status_filter == 'all':
            # Get all reports except drafts
            statuses = ['submitted', 'approved_leader', 'approved_admin', 'needs_revision']
            reports = []
            for status in statuses:
                reports.extend(get_reports_by_status(status))
        else:
            reports = get_reports_by_status(status_filter)
        
        # Clear the table
        self.all_reports_table.setRowCount(0)
        row = 0
        
        for report in reports:
            # Apply search filter
            if search_text and search_text not in report['title'].lower():
                continue
            
            self.all_reports_table.insertRow(row)
            
            # ID
            id_item = QTableWidgetItem(str(report['report_id']))
            self.all_reports_table.setItem(row, 0, id_item)
            
            # Title
            title_item = QTableWidgetItem(report['title'])
            self.all_reports_table.setItem(row, 1, title_item)
            
            # Created by
            creator_item = QTableWidgetItem(report['creator_name'])
            self.all_reports_table.setItem(row, 2, creator_item)
            
            # Submitted on
            created_item = QTableWidgetItem(report['created_at'])
            self.all_reports_table.setItem(row, 3, created_item)
            
            # Status with color coding
            status_text = self.get_status_display_text(report['status'])
            status_item = QTableWidgetItem(status_text)
            status_item.setForeground(self.get_status_color(report['status']))
            self.all_reports_table.setItem(row, 4, status_item)
            
            # Actions
            actions_cell = QWidget()
            actions_layout = QHBoxLayout(actions_cell)
            actions_layout.setContentsMargins(2, 2, 2, 2)
            
            # View button
            view_btn = QPushButton("View")
            view_btn.setStyleSheet("background-color: #3498db; color: white;")
            view_btn.clicked.connect(lambda checked, r=report['report_id']: self.view_report(r))
            actions_layout.addWidget(view_btn)
            
            # View PDF button
            pdf_btn = QPushButton("View PDF")
            pdf_btn.setStyleSheet("background-color: #8e44ad; color: white;")
            def open_pdf(report_id=report['report_id']):
                import os
                from PyQt5.QtWidgets import QMessageBox
                from PyQt5.QtGui import QDesktopServices
                from PyQt5.QtCore import QUrl
                # Try both approved_reports and admin_all_reports
                pdf_filename = f"report_{report_id}_"
                found_pdf = None
                for folder in ["approved_reports", "admin_all_reports"]:
                    dir_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), folder)
                    if os.path.isdir(dir_path):
                        for fname in os.listdir(dir_path):
                            if fname.startswith(pdf_filename) and fname.endswith(".pdf"):
                                found_pdf = os.path.join(dir_path, fname)
                                break
                    if found_pdf:
                        break
                if found_pdf and os.path.exists(found_pdf):
                    QDesktopServices.openUrl(QUrl.fromLocalFile(found_pdf))
                else:
                    QMessageBox.warning(self, "PDF Not Found", "No generated PDF found for this report.")
            pdf_btn.clicked.connect(lambda checked, r=report['report_id']: open_pdf(r))
            actions_layout.addWidget(pdf_btn)

            # Add approve/reject buttons only for reports approved by unit leader
            if report['status'] == 'approved_leader':
                # Approve button
                approve_btn = QPushButton("Approve")
                approve_btn.setStyleSheet("background-color: #2ecc71; color: white;")
                approve_btn.clicked.connect(lambda checked, r=report['report_id']: self.approve_report(r))
                actions_layout.addWidget(approve_btn)
                
                # Send Back button
                reject_btn = QPushButton("Send Back")
                reject_btn.setStyleSheet("background-color: #e74c3c; color: white;")
                reject_btn.clicked.connect(lambda checked, r=report['report_id']: self.send_back_report(r))
                actions_layout.addWidget(reject_btn)
            
            self.all_reports_table.setCellWidget(row, 5, actions_cell)
            
            row += 1
    
    def refresh_dashboard_stats(self):
        """Refresh the dashboard statistics"""
        # Get counts for different statuses
        pending_reports = get_reports_by_status('approved_leader')
        approved_reports = get_reports_by_status('approved_admin')
        users = get_all_users()
        
        # Update labels
        self.pending_count_label.setText(f"Pending Approvals: {len(pending_reports)}")
        self.users_count_label.setText(f"Users: {len(users)}")
        self.approved_count_label.setText(f"Approved Reports: {len(approved_reports)}")
    
    def add_user(self):
        """Add a new user"""
        try:
            dialog = UserManagementDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                user_data = dialog.get_user_data()
                
                # Validate data
                if not user_data['username'] or not user_data['password'] or not user_data['email']:
                    QMessageBox.warning(self, "Missing Data", "Username, password and email are required.")
                    return
                
                # Register the new user
                if Auth.register_user(
                    user_data['username'], 
                    user_data['password'], 
                    user_data['role'], 
                    user_data['email'],
                    user_data.get('emp_code', ''),
                    user_data.get('designation', '')
                ):
                    QMessageBox.information(self, "Success", "User created successfully.")
                    self.refresh_users_table()
                else:
                    QMessageBox.warning(self, "Error", "Failed to create user. Username might already exist.")
        except Exception as e:
            import logging
            logging.error(f"Error in add_user: {e}")
            QMessageBox.warning(self, "Error", f"An error occurred: {str(e)}")
            
    def edit_user(self, user):
        """Edit an existing user"""
        try:
            dialog = UserManagementDialog(self, user)
            if dialog.exec_() == QDialog.Accepted:
                user_data = dialog.get_user_data()
                
                # Validate data
                if not user_data['username'] or not user_data['email']:
                    QMessageBox.warning(self, "Missing Data", "Username and email are required.")
                    return
                
                # Update user in database
                password = user_data.get('password')
                password_hash = Auth.hash_password(password) if password else None
                success = update_user(
                    user['id'],
                    username=user_data.get('username'),
                    password_hash=password_hash,
                    role=user_data.get('role'),
                    email=user_data.get('email'),
                    emp_code=user_data.get('emp_code'),
                    designation=user_data.get('designation')
                )
                if success:
                    QMessageBox.information(self, "Success", "User updated successfully.")
                else:
                    QMessageBox.warning(self, "Error", "Failed to update user.")
                
                # Refresh the table
                self.refresh_users_table()
        except Exception as e:
            import logging
            logging.error(f"Error in edit_user: {e}")
            QMessageBox.warning(self, "Error", f"An error occurred: {str(e)}")
    
    def delete_user(self, user):
        """Delete a user"""
        # Don't allow deleting the current user or the default admin
        if user['id'] == self.user['id']:
            QMessageBox.warning(self, "Cannot Delete", "You cannot delete your own account.")
            return
            
        if user['username'] == 'admin':
            QMessageBox.warning(self, "Cannot Delete", "The default admin account cannot be deleted.")
            return
        
        # Confirm deletion
        reply = QMessageBox.question(
            self, "Confirm Deletion", 
            f"Are you sure you want to delete user {user['username']}?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Delete user from database
            success = delete_user(user['id'])
            
            if success:
                QMessageBox.information(self, "Success", f"User {user['username']} has been deleted.")
            else:
                QMessageBox.warning(self, "Error", "Failed to delete user.")
            
            # Refresh the table
            self.refresh_users_table()
    
    def upload_template(self):
        """Upload a new Excel template"""
        if not self.user['id']:
            QMessageBox.warning(self, "Error", "User ID not found.")
            return
        
        # Open file dialog to select Excel file
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel Template", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        # Get template name from file name
        template_name = os.path.basename(file_path)
        
        # Save the file to templates directory
        templates_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
        os.makedirs(templates_dir, exist_ok=True)
        
        # Copy file to templates directory with a new unique name
        import shutil
        import time
        
        new_file_name = f"template_{int(time.time())}_{template_name}"
        new_file_path = os.path.join(templates_dir, new_file_name)
        
        try:
            shutil.copy2(file_path, new_file_path)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to copy template file: {str(e)}")
            return
        
        # Add template to database
        template_id = add_template(template_name, new_file_path, self.user['id'])
        
        if template_id:
            QMessageBox.information(self, "Success", "Template uploaded successfully.")
            self.refresh_data()
            
            # Show preview of the template
            from utils.create_sample_template import create_sample_template
            if not os.path.exists(new_file_path):
                create_sample_template()
            
            self.preview_template(new_file_path)
        else:
            QMessageBox.warning(self, "Error", "Failed to upload template.")
    
    def preview_template(self, template_path):
        """Show a preview of a template"""
        if not os.path.exists(template_path):
            QMessageBox.warning(self, "Error", f"Template file not found: {template_path}")
            return
        
        dialog = TemplatePreviewDialog(template_path, self)
        dialog.exec_()
    
    def view_report(self, report_id):
        """View a report's full details including approval history"""
        report = get_report(report_id)
        if not report:
            QMessageBox.warning(self, "Error", "Report not found.")
            return
        
        # Create a more comprehensive ReportDetailDialog
        detail_dialog = ReportDetailDialog(report, self)
        detail_dialog.exec_()
    
    def approve_report(self, report_id):
        """Approve a report after review"""
        # Log to file to confirm approve_report is called
        with open("admin_approve_debug_log.txt", "a") as f:
            f.write(f"approve_report called for report_id={report_id} by user={self.user}\n")
        
        report = get_report(report_id)
        if not report:
            QMessageBox.warning(self, "Error", "Report not found.")
            return
        
        # Create confirmation dialog
        confirm_dialog = QDialog(self)
        confirm_dialog.setWindowTitle("Confirm Report Approval")
        confirm_dialog.setMinimumWidth(500)
        confirm_layout = QVBoxLayout(confirm_dialog)
        
        # Add header
        header = QLabel("<h3>Confirm Final Approval</h3>")
        header.setAlignment(Qt.AlignCenter)
        confirm_layout.addWidget(header)
        
        # Report summary section
        summary_group = QGroupBox("Report Summary")
        summary_layout = QFormLayout(summary_group)
        
        title = report.get('title', 'Untitled Report')
        summary_layout.addRow("Title:", QLabel(f"<b>{title}</b>"))
        
        author_name = "Unknown"
        if 'user_id' in report:
            author_name = get_user_name(report['user_id'])
        summary_layout.addRow("Author:", QLabel(author_name))
        
        if 'created_at' in report and report['created_at']:
            try:
                created_date = report['created_at'].strftime("%Y-%m-%d %H:%M")
            except (AttributeError, TypeError):
                created_date = str(report['created_at'])
            summary_layout.addRow("Created:", QLabel(created_date))
        
        if 'last_modified_at' in report and report['last_modified_at']:
            try:
                updated_date = report['last_modified_at'].strftime("%Y-%m-%d %H:%M")
            except (AttributeError, TypeError):
                updated_date = str(report['last_modified_at'])
            summary_layout.addRow("Last Updated:", QLabel(updated_date))
        
        # Show leader approval if applicable
        if report.get('status') == 'leader_approved':
            leader_approval = get_leader_approval(report_id)
            if leader_approval:
                leader_name = get_user_name(leader_approval.get('action_by', 0))
                
                # Format approval date safely
                leader_date = "Unknown date"
                if 'timestamp' in leader_approval and leader_approval['timestamp']:
                    try:
                        if isinstance(leader_approval['timestamp'], datetime):
                            leader_date = leader_approval['timestamp'].strftime("%Y-%m-%d %H:%M")
                        else:
                            leader_date = str(leader_approval['timestamp'])
                    except Exception:
                        leader_date = str(leader_approval['timestamp'])
                
                summary_layout.addRow("Unit Leader Approval:", 
                                    QLabel(f"Approved by {leader_name} on {leader_date}"))
        
        confirm_layout.addWidget(summary_group)
        
        # Comments section
        comments_group = QGroupBox("Final Approval Comments")
        comments_layout = QVBoxLayout(comments_group)
        
        comments_field = QTextEdit()
        comments_field.setPlaceholderText("Add any final comments about this approval (optional)")
        comments_field.setMaximumHeight(100)
        comments_layout.addWidget(comments_field)
        
        confirm_layout.addWidget(comments_group)
        
        # Warning message
        warning = QLabel("<i>Once approved, this report will be finalized and locked for editing.</i>")
        warning.setStyleSheet("color: #e74c3c;")
        warning.setAlignment(Qt.AlignCenter)
        confirm_layout.addWidget(warning)
        
        # Buttons
        button_layout = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(confirm_dialog.reject)
        
        approve_btn = QPushButton("Confirm Approval")
        approve_btn.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold;")
        approve_btn.clicked.connect(confirm_dialog.accept)
        
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(approve_btn)
        confirm_layout.addLayout(button_layout)
        
        # Show dialog and process result
        if confirm_dialog.exec_() == QDialog.Accepted:
            comments = comments_field.toPlainText()
            
            # Update report status and add approval log before PDF generation
            update_report_status(report_id, 'approved_admin', self.user['id'])
            admin_signature = f"{self.user.get('emp_code', '')} - {self.user.get('designation', '')}"
            log_comments = comments
            if admin_signature.strip():
                log_comments = f"{comments}\nDigital Signature: {admin_signature}" if comments else f"Digital Signature: {admin_signature}"
            add_approval_log(report_id, self.user['id'], 'approve_admin', log_comments)
            
            # Generate PDF with digital signatures
            try:
                # Get an active template for PDF generation
                active_template = get_active_template()
                with open("admin_approve_debug_log.txt", "a") as f:
                    f.write(f"active_template: {active_template}\n")
                if active_template:
                    # If active_template is a list, select the first template
                    if isinstance(active_template, list):
                        template_row = active_template[0]
                    else:
                        template_row = active_template
                    template_path = template_row['file_path']
                    with open("admin_approve_debug_log.txt", "a") as f:
                        f.write(f"template_path: {template_path}\n")
                    # Create PDF directory if it doesn't exist
                    pdf_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'approved_reports')
                    os.makedirs(pdf_dir, exist_ok=True)
                    
                    # Generate PDF filename
                    pdf_filename = f"report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                    pdf_path = os.path.join(pdf_dir, pdf_filename)
                    with open("admin_approve_debug_log.txt", "a") as f:
                        f.write(f"pdf_path: {pdf_path}\n")
                    
                    # Log to file just before PDF generation
                    with open("admin_approve_debug_log.txt", "a") as f:
                        f.write(f"Calling PDFGenerator.generate_report_pdf for report_id={report_id}\n")
                    
                    # Generate the PDF
                    success, result = PDFGenerator.generate_report_pdf(report_id, template_path, pdf_path)
                    
                    if success:
                        # Show success message with PDF path
                        QMessageBox.information(
                            self, 
                            "PDF Generated",
                            f"Report has been approved and PDF has been generated at:\n{pdf_path}"
                        )
                        # Open the PDF for preview
                        try:
                            if sys.platform == 'win32':
                                os.startfile(pdf_path)
                            elif sys.platform == 'darwin':  # macOS
                                subprocess.run(['open', pdf_path])
                            else:  # Linux
                                subprocess.run(['xdg-open', pdf_path])
                        except Exception as e:
                            print(f"Error opening PDF: {e}")
                    else:
                        QMessageBox.warning(
                            self, 
                            "PDF Generation Error",
                            f"Failed to generate PDF: {result}"
                        )
                else:
                    QMessageBox.warning(
                        self, 
                        "No Templates",
                        "No templates found for PDF generation."
                    )
            except Exception as e:
                QMessageBox.warning(self, "Error", f"An error occurred while generating the PDF: {str(e)}")
                
            # Send notification to report author
            try:
                # Use safe username from user dictionary
                user_name = self.user.get('username', 'Administrator')
                
                # Get the user ID for notifications
                user_id = report.get('user_id')
                if user_id:
                    send_report_notification(
                        user_id,
                        f"Report '{title}' Approved",
                        f"Your report has received final approval from {user_name}."
                    )
                
                # If there was a unit leader approval, notify them as well
                if report.get('status') == 'leader_approved':
                    leader_approval = get_leader_approval(report_id)
                    if leader_approval and 'action_by' in leader_approval:
                        send_report_notification(
                            leader_approval['action_by'],
                            f"Report '{title}' Final Approval",
                            f"A report you approved has received final approval from {user_name}."
                        )
                
            except Exception as e:
                QMessageBox.warning(
                    self, 
                    "Notification Error",
                    f"Report was approved but notification failed: {str(e)}"
                )
                
            # Refresh only the tables, keeping the current view
            self.refresh_pending_tables()
            self.refresh_all_reports_table()
        else:
            QMessageBox.warning(self, "Error", "Failed to approve report.")
    
    def send_back_report(self, report_id):
        """Send a report back for revision"""
        report = get_report(report_id)
        if not report:
            QMessageBox.warning(self, "Error", "Report not found.")
            return
        
        # Create dialog for revision comments
        dialog = QDialog(self)
        dialog.setWindowTitle("Send Report Back for Revision")
        dialog.setMinimumWidth(500)
        layout = QVBoxLayout(dialog)
        
        # Add header
        header = QLabel("<h3>Send Back for Revision</h3>")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        # Report info
        title = report.get('title', 'Untitled Report')
        info_label = QLabel(f"Report: <b>{title}</b>")
        layout.addWidget(info_label)
        
        # Comments section
        comments_label = QLabel("Please provide feedback for the author:")
        layout.addWidget(comments_label)
        
        comments_field = QTextEdit()
        comments_field.setPlaceholderText("Explain why the report needs revision and what changes are required...")
        layout.addWidget(comments_field)
        
        # Warning about required comments
        warning = QLabel("<i>Comments are required when sending a report back for revision.</i>")
        warning.setStyleSheet("color: #e74c3c;")
        layout.addWidget(warning)
        
        # Buttons
        button_layout = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
        
        send_btn = QPushButton("Send Back for Revision")
        send_btn.setStyleSheet("background-color: #e67e22; color: white;")
        
        def validate_and_submit():
            if not comments_field.toPlainText().strip():
                QMessageBox.warning(dialog, "Required Field", "Please provide revision comments for the author.")
                return
            dialog.accept()
            
        send_btn.clicked.connect(validate_and_submit)
        
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(send_btn)
        layout.addLayout(button_layout)
        
        # Process result
        if dialog.exec_() == QDialog.Accepted:
            comments = comments_field.toPlainText()
            
            # Update report status
            success = update_report_status(report_id, 'needs_revision', self.user['id'])
            if success:
                # Add approval log
                add_approval_log(report_id, self.user['id'], 'send_back', comments)
                
                # Send notification to report author
                try:
                    # Use safe user ID
                    user_id = report.get('user_id')
                    if user_id:
                        send_report_notification(
                            user_id,
                            f"Report '{title}' Needs Revision",
                            f"Your report requires revision. Comments: {comments}"
                        )
                except Exception as e:
                    QMessageBox.warning(
                        self, 
                        "Notification Error",
                        f"Report was sent back but notification failed: {str(e)}"
                    )
                
                QMessageBox.information(self, "Success", "Report has been sent back for revision.")
                self.refresh_data()
            else:
                QMessageBox.warning(self, "Error", "Failed to update report status.")
    
    def get_status_display_text(self, status):
        """Convert status code to display text"""
        status_map = {
            'draft': 'Draft',
            'submitted': 'Submitted',
            'approved_leader': 'Approved by Unit Leader',
            'approved_admin': 'Approved (Final)',
            'needs_revision': 'Needs Revision'
        }
        return status_map.get(status, status.capitalize())
    
    def get_status_color(self, status):
        """Get color for status"""
        status_colors = {
            'draft': QColor(128, 128, 128),  # Grey
            'submitted': QColor(52, 152, 219),  # Blue
            'approved_leader': QColor(241, 196, 15),  # Yellow
            'approved_admin': QColor(46, 204, 113),  # Green
            'needs_revision': QColor(231, 76, 60)  # Red
        }
        return status_colors.get(status, QColor(0, 0, 0))
    
    def refresh_templates_table(self):
        """Refresh the templates table"""
        templates = get_all_templates()
        
        # Clear the table
        self.templates_table.setRowCount(0)
        
        for row, template in enumerate(templates):
            self.templates_table.insertRow(row)
            
            # ID
            id_item = QTableWidgetItem(str(template['id']))
            self.templates_table.setItem(row, 0, id_item)
            
            # Name
            name_item = QTableWidgetItem(template['name'])
            self.templates_table.setItem(row, 1, name_item)
            
            # Uploaded By
            uploader = get_user_by_id(template['uploaded_by'])
            uploader_name = uploader['username'] if uploader else "Unknown"
            uploader_item = QTableWidgetItem(uploader_name)
            self.templates_table.setItem(row, 2, uploader_item)
            
            # Date
            date_item = QTableWidgetItem(str(template['uploaded_at']))
            self.templates_table.setItem(row, 3, date_item)
            
            # Actions
            actions_cell = QWidget()
            actions_layout = QHBoxLayout(actions_cell)
            actions_layout.setContentsMargins(2, 2, 2, 2)
            
            # Preview button
            preview_btn = QPushButton("Preview")
            preview_btn.setStyleSheet("background-color: #3498db; color: white;")
            preview_btn.clicked.connect(lambda checked, path=template['file_path']: self.preview_template(path))
            actions_layout.addWidget(preview_btn)
            
            # Delete button
            delete_btn = QPushButton("Delete")
            delete_btn.setStyleSheet("background-color: #e74c3c; color: white;")
            delete_btn.clicked.connect(lambda checked, t=template: self.delete_template(t['id']))
            actions_layout.addWidget(delete_btn)
            
            self.templates_table.setCellWidget(row, 4, actions_cell)
            
        # Resize columns for better view
        self.templates_table.resizeColumnsToContents()
        self.templates_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        
    def delete_template(self, template_id):
        """Delete a template from the system."""
        confirm_dialog = QMessageBox()
        confirm_dialog.setWindowTitle("Confirm Template Deletion")
        confirm_dialog.setText("Are you sure you want to delete this template?")
        confirm_dialog.setInformativeText("This action cannot be undone.")
        confirm_dialog.setIcon(QMessageBox.Warning)
        confirm_dialog.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        confirm_dialog.setDefaultButton(QMessageBox.No)
        
        if confirm_dialog.exec_() == QMessageBox.Yes:
            try:
                delete_template(template_id)
                QMessageBox.information(
                    self, 
                    "Success", 
                    "Template deleted successfully."
                )
                self.refresh_templates_table()
            except Exception as e:
                QMessageBox.critical(
                    self, 
                    "Error", 
                    f"Failed to delete template: {str(e)}"
                )
                logger.error(f"Error deleting template: {e}") 