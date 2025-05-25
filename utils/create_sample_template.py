import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def create_sample_template():
    """Create a sample Excel template with form fields for testing"""
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    
    # Define some styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "DAILY PRODUCTION REPORT"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add basic info section
    ws.append([])
    ws.append(["Report Information"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')
    cell = ws[f'A{ws.max_row}']
    cell.font = header_font
    
    # Add form fields
    ws.append(["Report Date:", "{{date}}", "", "Shift:", "{{shift}}"])
    ws.append(["Department:", "{{department}}", "", "Unit:", "{{unit}}"])
    ws.append(["Supervisor:", "{{supervisor}}", "", "Report ID:", "{{report_id}}"])
    
    # Add production section
    ws.append([])
    ws.append(["Production Data"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')
    cell = ws[f'A{ws.max_row}']
    cell.font = header_font
    
    # Add production table headers
    headers = ["Item", "Planned Quantity", "Actual Quantity", "Variance", "Comments"]
    ws.append(headers)
    
    # Format headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add production items
    items = ["Product A", "Product B", "Product C", "Product D"]
    for item in items:
        field_base = item.lower().replace(" ", "_")
        row = [
            item, 
            f"{{{{planned_{field_base}}}}}", 
            f"{{{{actual_{field_base}}}}}", 
            f"={{{{actual_{field_base}}}}} - {{{{planned_{field_base}}}}}", 
            f"{{{{comments_{field_base}}}}}"
        ]
        ws.append(row)
        
        # Apply borders to all cells in the row
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).border = border
    
    # Add issues section
    ws.append([])
    ws.append(["Issues and Challenges"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')
    cell = ws[f'A{ws.max_row}']
    cell.font = header_font
    
    # Add issues table
    headers = ["Issue Description", "Impact", "Action Taken", "Status"]
    ws.append(headers)
    
    # Format headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add rows for issues
    for i in range(1, 4):
        row = [
            f"{{{{issue_{i}_description}}}}", 
            f"{{{{issue_{i}_impact}}}}", 
            f"{{{{issue_{i}_action}}}}", 
            f"{{{{issue_{i}_status}}}}"
        ]
        ws.append(row)
        
        # Apply borders to all cells in the row
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).border = border
    
    # Add notes section
    ws.append([])
    ws.append(["Additional Notes"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')
    cell = ws[f'A{ws.max_row}']
    cell.font = header_font
    
    ws.append(["{{notes}}"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row+2}')
    cell = ws[f'A{ws.max_row}']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = border
    
    # Save the template
    template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    template_path = os.path.join(template_dir, 'sample_template.xlsx')
    wb.save(template_path)
    
    print(f"Sample template created at: {template_path}")
    return template_path

if __name__ == "__main__":
    create_sample_template() 