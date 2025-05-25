import sys
import os
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Add the parent directory to the path to allow imports
parent_dir = str(Path(__file__).resolve().parent.parent)
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

class ExcelHandler:
    @staticmethod
    def get_template_fields(template_path):
        """Extract fields from an Excel template to create a dynamic form"""
        try:
            workbook = openpyxl.load_workbook(template_path, data_only=True)
            sheet = workbook.active
            
            # Create a list to store fields
            fields = []
            
            # Iterate through the sheet and look for input cells
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    # Check if this is a field that needs user input
                    # We'll look for cells with "{{field_name}}" pattern
                    cell_value = str(cell.value) if cell.value else ""
                    if "{{" in cell_value and "}}" in cell_value:
                        # Extract field name from {{field_name}}
                        field_name = cell_value.replace("{{", "").replace("}}", "").strip()
                        
                        # Determine field type based on field name
                        field_type = "text"  # Default
                        if "date" in field_name.lower():
                            field_type = "date"
                        elif "number" in field_name.lower() or "quantity" in field_name.lower():
                            field_type = "number"
                        
                        fields.append({
                            'name': field_name,
                            'type': field_type,
                            'row': row,
                            'col': col,
                            'col_letter': get_column_letter(col)
                        })
            
            return fields
        except Exception as e:
            print(f"Error reading template: {str(e)}")
            return []
    
    @staticmethod
    def create_report_from_template(template_path, output_path, data):
        """
        Fill in a template with report data and save to output_path
        data: Dictionary with keys matching field names and values to insert
        """
        # Defensive check
        if not isinstance(data, dict):
            print(f"ERROR: Expected dict for data, got {type(data)} with value {data}")
            import traceback
            traceback.print_stack()
            return False
        print(f"DEBUG: create_report_from_template using data: {data}")
        try:
            # Load the template
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook.active
            
            # Fill in the template with data
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = str(cell.value) if cell.value else ""
                    
                    # Check if this is a field that needs user input
                    if "{{" in cell_value and "}}" in cell_value:
                        field_name = cell_value.replace("{{", "").replace("}}", "").strip()
                        
                        # If we have data for this field, replace it
                        if field_name in data:
                            cell.value = data[field_name]
            
            # Add signature information if provided
            if 'signatures' in data:
                signatures = data['signatures']
                row = sheet.max_row + 2  # Add some space
                
                sheet.cell(row=row, column=1).value = "Signatures:"
                row += 1
                
                for sig in signatures:
                    sheet.cell(row=row, column=1).value = f"{sig['role']}:"
                    sheet.cell(row=row, column=2).value = sig['name']
                    sheet.cell(row=row, column=3).value = sig['timestamp']
                    row += 1
            
            # Save the filled template
            workbook.save(output_path)
            return True
        except Exception as e:
            print(f"Error creating report from template: {str(e)}")
            return False
    
    @staticmethod
    def get_template_preview(template_path, max_rows=10):
        """Get a preview of the template for display in the UI"""
        try:
            workbook = openpyxl.load_workbook(template_path, data_only=True)
            sheet = workbook.active
            
            preview = []
            for row in range(1, min(max_rows + 1, sheet.max_row + 1)):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                preview.append(row_data)
            
            return preview
        except Exception as e:
            print(f"Error getting template preview: {str(e)}")
            return [] 